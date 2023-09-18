'''
Takes a xlmx file of the attacks as an input and converts the attack timestamps into ecpoch time 
'''
import json
import openpyxl
from datetime import datetime,timedelta

def convert_to_epoch(date_string):
    print(date_string, type(date_string))
    epoch = datetime(1970, 1, 1)
    delta = date_string - epoch
    return delta.total_seconds()

def generateJson(name, list):
    try:
        with open(f"{output_path}{name}.json", "w") as file:
            json.dump(list, file, indent=4)
        print(f"List successfully saved as JSON to {name}.json")
    except Exception as e:
        print(f"An error occurred: {e}")


def process_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            attack_timeframes = []
            if sheet_name != "Foglio1": # skip overview sheet
                sheet = workbook[sheet_name]
                id = 1
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming header is in the first row
                    start_time = row[1]  # Assuming "Start time" is in the second column
                    end_time = row[2]
                    if row[1]== None:
                        break
                    if sheet_name == 'attack_4':
                        start_time -= timedelta(hours=1)
                        end_time -= timedelta(hours=1)
                    else:
                        start_time -= timedelta(hours=2)
                        end_time -= timedelta(hours=2)
                    epoch_starttime = convert_to_epoch(start_time)
                    epoch_endtime = convert_to_epoch(end_time)
                    interval_entry = {
                        "id": id,
                        "description": "",
                        "start": float(epoch_starttime),
                        "end": float(epoch_endtime)
                    }
                    attack_timeframes.append(interval_entry)
                    print(id)
                    id += 1
                generateJson(sheet_name, attack_timeframes)
        
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    input_file = "attack/attacks.xlsx"  # Passe den Dateinamen an
    output_path = "attack/"
    process_excel_file(input_file)
